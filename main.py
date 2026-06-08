import os
import io
import json
import sqlite3
import base64
from openai import OpenAI
import asyncio
from pathlib import Path
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor

from fastapi import FastAPI, File, UploadFile, HTTPException
from fastapi.staticfiles import StaticFiles
from fastapi.responses import HTMLResponse, StreamingResponse
from dotenv import load_dotenv
import pdfplumber
import fitz  # PyMuPDF
from PIL import Image
from groq import Groq
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

try:
    import pytesseract
    TESSERACT_AVAILABLE = True
except ImportError:
    TESSERACT_AVAILABLE = False

load_dotenv()

app = FastAPI(title="Samtel Avionics Invoice Processor")

GROQ_API_KEY = os.getenv("GROQ_API_KEY")
if not GROQ_API_KEY:
    raise RuntimeError("GROQ_API_KEY not set in environment or .env file")

groq_client = Groq(api_key=GROQ_API_KEY)

OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")
openai_client = OpenAI(api_key=OPENAI_API_KEY) if OPENAI_API_KEY else None

ALLOWED_TYPES = {
    "application/pdf",
    "image/png",
    "image/jpeg",
    "image/jpg",
    "image/tiff",
    "image/bmp",
    "image/webp",
}

_executor = ThreadPoolExecutor(max_workers=4)

# ── Database ───────────────────────────────────────────────────────────────────

DB_PATH = Path(__file__).parent / "invoices.db"


def get_db_connection():
    con = sqlite3.connect(DB_PATH, timeout=30.0)
    # Enable Write-Ahead Logging (WAL) to allow concurrent reads and prevent locks
    con.execute("PRAGMA journal_mode=WAL;")
    return con


def _init_db():
    con = get_db_connection()
    con.execute("""CREATE TABLE IF NOT EXISTS invoices (
        id             INTEGER PRIMARY KEY AUTOINCREMENT,
        processed_at   TEXT NOT NULL,
        filename       TEXT,
        invoice_type   TEXT,
        invoice_number TEXT,
        invoice_date   TEXT,
        party_name     TEXT,
        total_amount   TEXT,
        data_json      TEXT,
        duplicate_of   INTEGER
    )""")
    con.commit()
    con.close()


_init_db()


def _extract_summary(invoice_type: str, data: dict) -> tuple:
    num   = str(data.get("invoice_number") or "")
    date  = str(data.get("invoice_date") or "")
    party = str(
        data.get("vendor_name") or data.get("service_provider_name") or
        data.get("courier_company") or data.get("seller_name") or
        data.get("employee_name") or ""
    )
    total = str(
        data.get("grand_total") or data.get("total_inr") or
        data.get("gross_total") or data.get("total_amount") or
        data.get("net_payable_to_employee") or ""
    )
    return num, date, party, total


def _check_duplicate(invoice_number: str, invoice_type: str):
    if not invoice_number or invoice_number in ("None", ""):
        return None
    con = get_db_connection()
    row = con.execute(
        "SELECT id, filename FROM invoices WHERE invoice_number=? AND invoice_type=? "
        "AND duplicate_of IS NULL ORDER BY processed_at DESC LIMIT 1",
        (invoice_number, invoice_type),
    ).fetchone()
    con.close()
    return row


def _save_invoice(filename, invoice_type, data, duplicate_of=None) -> int:
    num, date, party, total = _extract_summary(invoice_type, data)
    con = get_db_connection()
    cur = con.execute(
        "INSERT INTO invoices (processed_at,filename,invoice_type,invoice_number,"
        "invoice_date,party_name,total_amount,data_json,duplicate_of) VALUES (?,?,?,?,?,?,?,?,?)",
        (
            datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%S"),
            filename, invoice_type, num, date, party, total,
            json.dumps(data), duplicate_of,
        ),
    )
    row_id = cur.lastrowid
    con.commit()
    con.close()
    return row_id


# ── Invoice type classifier ────────────────────────────────────────────────────

CLASSIFY_PROMPT = """You are an invoice classifier for Samtel Avionics Limited, an avionics company in India.

STEP 1 — Find the ISSUER (the company that printed/sent the invoice):
- Look for the company name at the very top of the document, or in the "From" / "Supplier" / "Vendor" field.
- Samtel Avionics / Samtel HAL Display Systems appearing as ISSUER → SALE_DOMESTIC or SALE_EXPORT.
- Any other Indian company appearing as ISSUER, with Samtel as the buyer/billed-to → PURCHASE_DOMESTIC or SERVICES.
- UPS / DHL / FedEx as ISSUER with IEC code / foreign shipment → IMPORT.
- Government e-Marketplace (GEM) as ISSUER → SERVICES.
- Reimbursement/expense form with no issuer company → EMPLOYEE_REIMBURSEMENT.

STEP 2 — Pick exactly one code:

PURCHASE_DOMESTIC   - A non-Samtel Indian vendor issued the invoice; Samtel is the buyer. Goods/materials purchased.
IMPORT              - UPS/DHL/courier issued it; involves foreign shipment, IEC code, duty, foreign currency.
SALE_DOMESTIC       - Samtel Avionics (or Samtel HAL) issued the invoice; Indian buyer.
SALE_EXPORT         - Samtel issued the invoice; buyer is outside India (foreign address, no GST, SWIFT).
SERVICES            - A service/job-work provider OR GEM issued the invoice for services rendered to Samtel.
                      STRONG signals: document title says "JOB WORK INVOICE", HSN/SAC code starts with 99 (e.g. 998898),
                      description contains words like Manufacturing, Assembly, PCBA, Processing, Job Work, GEM, Transaction charges.
EMPLOYEE_REIMBURSEMENT - Internal employee expense reimbursement form.

Return ONLY the type code, nothing else.

Invoice text:
{invoice_text}"""

# ── Type-specific extraction prompts ──────────────────────────────────────────

PROMPTS = {

"PURCHASE_DOMESTIC": """Extract ALL fields from this domestic purchase invoice and return valid JSON only.

Fields:
- invoice_number
- invoice_date
- irn_number
- ack_number
- ack_date
- eway_bill_number
- vendor_name
- vendor_gstin
- vendor_pan
- vendor_address
- vendor_state
- vendor_state_code
- bill_to_name
- bill_to_address
- bill_to_gstin
- ship_to_name
- ship_to_address
- ship_to_gstin
- po_number
- po_date
- payment_terms
- line_items: array of {{ sno, material_code, description, hsn_sac, quantity, uom, rate, total, discount_pct, taxable_value, cgst_pct, cgst_amt, sgst_pct, sgst_amt, igst_pct, igst_amt }}
- taxable_total
- total_cgst
- total_sgst
- total_igst
- total_gst
- grand_total
- amount_in_words
- currency
- vendor_bank: {{ beneficiary, bank_name, account_no, ifsc, swift }}

Return null for any missing field. Return ONLY valid JSON.

Invoice text:
{invoice_text}""",

"IMPORT": """Extract ALL fields from this import/courier invoice and return valid JSON only.

Fields:
- invoice_subtype  (BILL_OF_SUPPLY or TAX_INVOICE)
- courier_company
- account_number
- customer_iec_code
- customer_pan
- customer_gstin
- invoice_number
- invoice_date
- irn_number
- ack_number
- ack_date
- place_of_supply
- shipment: {{ import_date, tracking_number, reference_number, shipment_number, service_type, packages, bill_type, weight_kg }}
- goods: {{ description, value, original_currency, customs_number, exchange_rate_to_inr }}
- shipper_name
- shipper_location
- hsn_code
- duty_amount_inr
- disbursement_fee_inr
- sgst_pct
- sgst_amt
- cgst_pct
- cgst_amt
- igst_pct
- igst_amt
- total_inr
- courier_bank: {{ bank_name, account_number, ifsc, swift }}

Return null for any missing field. Return ONLY valid JSON.

Invoice text:
{invoice_text}""",

"SALE_DOMESTIC": """Extract ALL fields from this domestic sales invoice (Samtel Avionics is the SELLER) and return valid JSON only.

Fields:
- invoice_number
- invoice_date
- irn_number
- ack_number
- ack_date
- eway_bill_number
- seller_name
- seller_address
- seller_gstin
- seller_pan
- seller_cin
- seller_state
- seller_state_code
- buyer_name  (bill-to)
- buyer_address
- buyer_gstin
- buyer_state
- buyer_state_code
- consignee_name  (ship-to, may differ from buyer)
- consignee_address
- consignee_gstin
- consignee_state
- delivery_note_number
- samtel_reference
- buyer_order_number
- buyer_order_date
- payment_mode
- insurance_policy
- dispatched_through
- destination
- line_items: array of {{ sno, description, part_number, serial_numbers, hsn_sac, quantity, uom, rate, amount }}
- tax_base_amount
- igst_pct
- igst_amount
- cgst_pct
- cgst_amount
- sgst_pct
- sgst_amount
- total_amount
- amount_in_words
- payment_terms
- currency
- seller_bank: {{ ac_holder, bank_name, account_no, branch_ifsc, swift_code }}

Return null for any missing field. Return ONLY valid JSON.

Invoice text:
{invoice_text}""",

"SALE_EXPORT": """Extract ALL fields from this export sales invoice (Samtel Avionics is the SELLER, buyer is a foreign entity) and return valid JSON only.

Fields:
- invoice_number
- invoice_date
- irn_number  (may be blank for exports)
- ack_number
- ack_date
- eway_bill_number
- seller_name
- seller_address
- seller_gstin
- seller_pan
- seller_cin
- buyer_name
- buyer_address
- buyer_country
- buyer_gstin  (null for foreign buyers)
- consignee_name
- consignee_address
- consignee_country
- delivery_note_number
- samtel_reference
- buyer_order_number
- buyer_order_date
- payment_mode
- dispatched_through
- destination_country
- line_items: array of {{ sno, description, part_number, serial_numbers, hsn_sac, quantity, uom, rate_inr, amount_inr }}
- tax_base_amount_inr
- gst_applicable  (false for exports)
- igst_amount  (should be 0 for exports)
- total_amount_inr
- amount_in_words
- payment_terms
- currency
- seller_bank: {{ ac_holder, bank_name, account_no, branch_ifsc, swift_code }}

Return null for any missing field. Return ONLY valid JSON.

Invoice text:
{invoice_text}""",

"SERVICES": """Extract ALL fields from this service/job-work/GEM invoice and return valid JSON only.

First identify the sub-type: JOB_WORK, GEM, or GENERAL_SERVICE.

Fields:
- invoice_subtype  (JOB_WORK / GEM / GENERAL_SERVICE)
- invoice_number
- invoice_date
- irn_number
- ack_number
- ack_date
- service_provider_name
- service_provider_gstin
- service_provider_address
- service_provider_state
- bill_to_name
- bill_to_address
- bill_to_gstin
- bill_to_state
- place_of_supply
- rcm_applicable
- payment_terms
- line_items: array of {{ sno, description, order_number, hsn_sac, quantity, uom, rate, taxable_amount, gst_pct, gst_amount }}
- taxable_total
- cgst_pct
- cgst_amt
- sgst_pct
- sgst_amt
- igst_pct
- igst_amt
- total_gst
- gross_total
- tds_pct  (if TDS deducted, e.g. GEM invoices)
- tds_amount
- net_payable_after_tds
- amount_in_words
- currency
- nrgp_number  (for job-work)
- eway_bill_number
- challan_reference
- bank_details: {{ beneficiary, bank_name, account_no, ifsc }}

Return null for any missing field. Return ONLY valid JSON.

Invoice text:
{invoice_text}""",

"EMPLOYEE_REIMBURSEMENT": """Extract ALL fields from this employee reimbursement form and return valid JSON only.

Fields:
- employee_name
- employee_sap_code
- department
- budget_code
- advance_amount
- section_a_travel: {{ boarding, lodging, personal_allowance, total }}
- section_b_allowances: {{ total }}
- total_a_plus_b
- section_c_expenses: {{ conveyance, telephone, postage, stationery, misc_expenses, ticket_charges, business_promotion, total }}
- grand_total
- less_air_ticket_by_company
- net_payable_to_employee
- checked_by
- passed_by
- head_of_department
- currency

Return null for any missing field. Return ONLY valid JSON.

Invoice text:
{invoice_text}""",
}

# ── Text extraction ────────────────────────────────────────────────────────────

def _ocr_pdf_with_pymupdf(file_bytes: bytes, dpi: int = 300) -> str:
    if not TESSERACT_AVAILABLE:
        raise HTTPException(
            status_code=501,
            detail="This PDF appears to be scanned. OCR requires Tesseract: brew install tesseract",
        )
    doc = fitz.open(stream=file_bytes, filetype="pdf")
    parts = []
    mat = fitz.Matrix(dpi / 72, dpi / 72)
    for page in doc:
        pix = page.get_pixmap(matrix=mat, colorspace=fitz.csGRAY)
        img = Image.frombytes("L", (pix.width, pix.height), pix.samples)
        parts.append(pytesseract.image_to_string(img))
    doc.close()
    return "\n\n".join(parts)


def _ocr_image_with_tesseract(file_bytes: bytes) -> str:
    if not TESSERACT_AVAILABLE:
        raise HTTPException(
            status_code=501,
            detail="Image OCR requires Tesseract. Install it with: brew install tesseract",
        )
    image = Image.open(io.BytesIO(file_bytes))
    return pytesseract.image_to_string(image)


def _ocr_with_openai_vlm(file_bytes: bytes, is_pdf: bool) -> str:
    images = []
    if is_pdf:
        doc = fitz.open(stream=file_bytes, filetype="pdf")
        for page in doc:
            pix = page.get_pixmap(dpi=150)
            png_bytes = pix.tobytes("png")
            images.append((png_bytes, "image/png"))
        doc.close()
    else:
        try:
            img = Image.open(io.BytesIO(file_bytes))
            fmt = (img.format or "PNG").lower()
            mime = f"image/{fmt}"
        except Exception:
            mime = "image/png"
        images.append((file_bytes, mime))

    transcriptions = []
    for idx, (img_bytes, mime) in enumerate(images):
        b64_img = base64.b64encode(img_bytes).decode("utf-8")
        prompt = "You are an expert OCR engine. Transcribe all readable text from this invoice page. Do not summarize or omit any details. Maintain layout/tables where possible."
        if len(images) > 1:
            prompt = f"You are an expert OCR engine. Transcribe all readable text from page {idx + 1} of this invoice. Do not summarize or omit any details. Maintain layout/tables where possible."
            
        resp = openai_client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {
                    "role": "user",
                    "content": [
                        {"type": "text", "text": prompt},
                        {
                            "type": "image_url",
                            "image_url": {
                                "url": f"data:{mime};base64,{b64_img}"
                            }
                        }
                    ]
                }
            ],
            temperature=0.0,
        )
        text = resp.choices[0].message.content.strip()
        transcriptions.append(text)

    return "\n\n".join(transcriptions)


def extract_text_from_pdf(file_bytes: bytes) -> tuple[str, str]:
    text_parts = []
    with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if text:
                text_parts.append(text)
    if text_parts:
        return "\n\n".join(text_parts), "Digital (Native)"

    if openai_client:
        try:
            return _ocr_with_openai_vlm(file_bytes, is_pdf=True), "Scanned (VLM)"
        except Exception:
            return _ocr_pdf_with_pymupdf(file_bytes, dpi=300), "Scanned (OCR)"
    return _ocr_pdf_with_pymupdf(file_bytes, dpi=300), "Scanned (OCR)"


def extract_text_from_image(file_bytes: bytes) -> tuple[str, str]:
    if openai_client:
        try:
            return _ocr_with_openai_vlm(file_bytes, is_pdf=False), "Scanned (VLM)"
        except Exception:
            return _ocr_image_with_tesseract(file_bytes), "Scanned (OCR)"
    return _ocr_image_with_tesseract(file_bytes), "Scanned (OCR)"




# ── LLM helpers ───────────────────────────────────────────────────────────────

def _llm(prompt: str, max_tokens: int = 256) -> str:
    resp = groq_client.chat.completions.create(
        model="llama-3.3-70b-versatile",
        messages=[{"role": "user", "content": prompt}],
        temperature=0.1,
        max_tokens=max_tokens,
    )
    return resp.choices[0].message.content.strip()


def _parse_json(content: str) -> dict:
    if content.startswith("```"):
        lines = content.split("\n")
        content = "\n".join(lines[1:-1] if lines[-1].strip() == "```" else lines[1:])
    try:
        return json.loads(content)
    except json.JSONDecodeError:
        return {"raw_response": content, "parse_error": "Could not parse JSON from model response"}


def _keyword_classify(text: str) -> str | None:
    import re
    t = text.upper()

    if "JOB WORK INVOICE" in t or "JOB WORK" in t:
        return "SERVICES"
    if "GOVERNMENT E-MARKETPLACE" in t or "GOVERNMENT EMARKETPLACE" in t:
        return "SERVICES"
    if "GEM-20" in t and ("TRANSACTION CHARGES" in t or "MILESTONE" in t):
        return "SERVICES"
    if re.search(r'\b998[0-9]\d\d\b', t):
        return "SERVICES"

    if ("REIMBURSEMENT" in t or "T.A. BILL" in t or "TA BILL" in t or "TRAVEL ALLOWANCE" in t) and ("EMPLOYEE" in t or "SAP" in t or "DEPARTMENT" in t or "VOUCHER" in t):
        return "EMPLOYEE_REIMBURSEMENT"

    if re.search(r'IEC\s*(NO|CODE|NUMBER)?[\s:]*\d', t):
        return "IMPORT"
    if ("UPS EXPRESS" in t or "DHL EXPRESS" in t or "FEDEX" in t) and ("DUTY" in t or "USD" in t):
        return "IMPORT"

    address_cut = re.search(r'\b(SHIP\s*TO|BILL\s*TO|CONSIGNEE|BUYER|BILLED\s*TO|RECEIVER)\b', t)
    header_zone = t[:address_cut.start()] if address_cut else t[:400]
    samtel_is_issuer = "SAMTEL AVIONICS" in header_zone or "SAMTEL HAL DISPLAY" in header_zone

    if samtel_is_issuer:
        buyer_block = re.search(r'(BUYER|BILL\s*TO).{0,600}GSTIN[/\s]*UIN\s*:\s*(\S*)', t, re.DOTALL)
        if buyer_block:
            gstin_val = buyer_block.group(2).strip()
            if not gstin_val or len(gstin_val) < 10:
                return "SALE_EXPORT"
        return "SALE_DOMESTIC"

    return None


def classify_invoice(raw_text: str) -> str:
    fast = _keyword_classify(raw_text)
    if fast:
        return fast
    result = _llm(CLASSIFY_PROMPT.format(invoice_text=raw_text[:4000]), max_tokens=16)
    valid = {"PURCHASE_DOMESTIC", "IMPORT", "SALE_DOMESTIC", "SALE_EXPORT", "SERVICES", "EMPLOYEE_REIMBURSEMENT"}
    for v in valid:
        if v in result.upper():
            return v
    return "PURCHASE_DOMESTIC"


def extract_invoice_data(invoice_type: str, raw_text: str) -> dict:
    prompt = PROMPTS[invoice_type].format(invoice_text=raw_text)
    content = _llm(prompt, max_tokens=3000)
    return _parse_json(content)


def _extract_with_openai_vlm_direct(file_bytes: bytes, is_pdf: bool, invoice_type: str) -> dict:
    page_images = []
    if is_pdf:
        doc = fitz.open(stream=file_bytes, filetype="pdf")
        for page in doc:
            pix = page.get_pixmap(dpi=150)
            png_bytes = pix.tobytes("png")
            page_images.append(png_bytes)
        doc.close()
    else:
        page_images.append(file_bytes)

    if not page_images:
        raise ValueError("No pages found in file")

    content_list = [{"type": "text", "text": PROMPTS[invoice_type]}]
    for img_bytes in page_images:
        try:
            img = Image.open(io.BytesIO(img_bytes))
            fmt = (img.format or "PNG").lower()
            mime = f"image/{fmt}"
        except Exception:
            mime = "image/png"
        
        b64_img = base64.b64encode(img_bytes).decode("utf-8")
        content_list.append({
            "type": "image_url",
            "image_url": {
                "url": f"data:{mime};base64,{b64_img}"
            }
        })

    resp = openai_client.chat.completions.create(
        model="gpt-4o",
        messages=[
            {
                "role": "user",
                "content": content_list
            }
        ],
        response_format={"type": "json_object"},
        temperature=0.0,
    )
    result = resp.choices[0].message.content.strip()
    return _parse_json(result)


# ── Core processing helper ─────────────────────────────────────────────────────

def _process_file_sync(file_bytes: bytes, filename: str, content_type: str) -> dict:
    if content_type == "application/pdf":
        raw_text, ocr_method = extract_text_from_pdf(file_bytes)
    else:
        raw_text, ocr_method = extract_text_from_image(file_bytes)

    if not raw_text or len(raw_text.strip()) < 20:
        raise ValueError("Could not extract readable text. Ensure the file is not a low-resolution scan.")

    invoice_type  = classify_invoice(raw_text)
    
    # Direct VLM visual extraction ONLY for scanned EMPLOYEE_REIMBURSEMENT documents/images
    if invoice_type == "EMPLOYEE_REIMBURSEMENT" and ocr_method in ("Scanned (VLM)", "Scanned (OCR)") and openai_client:
        try:
            invoice_data = _extract_with_openai_vlm_direct(file_bytes, content_type == "application/pdf", invoice_type)
            ocr_method = f"{ocr_method} + Direct VLM"
        except Exception as e:
            print(f"Direct VLM extraction failed, falling back to Llama: {e}")
            invoice_data = extract_invoice_data(invoice_type, raw_text)
    else:
        invoice_data = extract_invoice_data(invoice_type, raw_text)

    inv_num = invoice_data.get("invoice_number")
    dup     = _check_duplicate(str(inv_num or ""), invoice_type)
    dup_warning = f"Possible duplicate of '{dup[1]}' (ID #{dup[0]})" if dup else None

    row_id = _save_invoice(filename, invoice_type, invoice_data, dup[0] if dup else None)

    # Save processed file to uploads directory
    try:
        ext = Path(filename).suffix or ".pdf"
        uploads_dir = Path(__file__).parent / "uploads"
        uploads_dir.mkdir(exist_ok=True)
        saved_file = uploads_dir / f"{row_id}{ext}"
        saved_file.write_bytes(file_bytes)
        file_url = f"/uploads/{row_id}{ext}"
    except Exception as e:
        print(f"Failed to save file to uploads: {e}")
        file_url = None

    return {
        "status":            "success",
        "id":                row_id,
        "filename":          filename,
        "invoice_type":      invoice_type,
        "ocr_method":        ocr_method,
        "file_url":          file_url,
        "invoice_data":      invoice_data,
        "duplicate_warning": dup_warning,
        "raw_text_preview":  raw_text[:500] + ("..." if len(raw_text) > 500 else ""),
    }



# ── API ────────────────────────────────────────────────────────────────────────

@app.post("/api/process-invoice")
async def process_invoice(file: UploadFile = File(...)):
    if file.content_type not in ALLOWED_TYPES:
        raise HTTPException(
            status_code=400,
            detail=f"Unsupported file type: {file.content_type}. Upload a PDF or image file.",
        )
    file_bytes = await file.read()
    if not file_bytes:
        raise HTTPException(status_code=400, detail="Uploaded file is empty.")
    try:
        return _process_file_sync(file_bytes, file.filename, file.content_type)
    except ValueError as e:
        raise HTTPException(status_code=422, detail=str(e))


@app.post("/api/process-batch")
async def process_batch(files: list[UploadFile] = File(...)):
    if len(files) > 20:
        raise HTTPException(400, "Maximum 20 files per batch.")

    sem  = asyncio.Semaphore(3)
    loop = asyncio.get_event_loop()

    async def _one(fb, fname, ctype):
        async with sem:
            try:
                return await loop.run_in_executor(
                    _executor, _process_file_sync, fb, fname, ctype
                )
            except Exception as e:
                return {"status": "error", "filename": fname, "error": str(e)}

    items = []
    for f in files:
        if f.content_type not in ALLOWED_TYPES:
            items.append({"error": f"Unsupported: {f.content_type}", "bytes": None,
                          "name": f.filename, "ctype": f.content_type})
        else:
            fb = await f.read()
            items.append({"bytes": fb, "name": f.filename, "ctype": f.content_type, "error": None})

    tasks = [
        (_one(it["bytes"], it["name"], it["ctype"]) if not it["error"]
         else asyncio.coroutine(lambda it=it: {"status": "error", "filename": it["name"], "error": it["error"]})())
        for it in items
    ]
    results = await asyncio.gather(*tasks, return_exceptions=True)

    return [
        r if not isinstance(r, Exception)
        else {"status": "error", "filename": items[i]["name"], "error": str(r)}
        for i, r in enumerate(results)
    ]


# ── History ────────────────────────────────────────────────────────────────────

@app.get("/api/history")
async def get_history():
    con = get_db_connection()
    rows = con.execute(
        "SELECT id,processed_at,filename,invoice_type,invoice_number,"
        "invoice_date,party_name,total_amount,duplicate_of FROM invoices "
        "ORDER BY processed_at DESC LIMIT 500"
    ).fetchall()
    con.close()
    keys = ["id","processed_at","filename","invoice_type","invoice_number",
            "invoice_date","party_name","total_amount","duplicate_of"]
    return [dict(zip(keys, r)) for r in rows]


@app.get("/api/history/{invoice_id}")
async def get_invoice(invoice_id: int):
    con = get_db_connection()
    row = con.execute("SELECT * FROM invoices WHERE id=?", (invoice_id,)).fetchone()
    con.close()
    if not row:
        raise HTTPException(404, "Invoice not found")
    keys = ["id","processed_at","filename","invoice_type","invoice_number",
            "invoice_date","party_name","total_amount","data_json","duplicate_of"]
    d = dict(zip(keys, row))
    d["invoice_data"] = json.loads(d.pop("data_json"))

    # Find matching saved file
    file_url = None
    uploads_dir = Path(__file__).parent / "uploads"
    if uploads_dir.exists():
        for f in uploads_dir.iterdir():
            if f.is_file() and f.name.startswith(f"{invoice_id}."):
                file_url = f"/uploads/{f.name}"
                break
    d["file_url"] = file_url
    return d



@app.delete("/api/history")
async def clear_history():
    con = get_db_connection()
    con.execute("DELETE FROM invoices")
    con.commit()
    con.close()
    return {"status": "cleared"}


# ── Analytics ──────────────────────────────────────────────────────────────────

@app.get("/api/analytics")
async def get_analytics():
    con = get_db_connection()
    total  = con.execute("SELECT COUNT(*) FROM invoices").fetchone()[0]
    dups   = con.execute("SELECT COUNT(*) FROM invoices WHERE duplicate_of IS NOT NULL").fetchone()[0]
    by_type = con.execute(
        "SELECT invoice_type, COUNT(*) FROM invoices GROUP BY invoice_type"
    ).fetchall()
    top_vendors = con.execute(
        "SELECT party_name, COUNT(*) FROM invoices WHERE party_name!='' "
        "GROUP BY party_name ORDER BY COUNT(*) DESC LIMIT 10"
    ).fetchall()
    daily = con.execute(
        "SELECT DATE(processed_at) as d, COUNT(*) FROM invoices "
        "GROUP BY d ORDER BY d ASC LIMIT 30"
    ).fetchall()
    con.close()
    return {
        "total_count":     total,
        "duplicate_count": dups,
        "by_type":    [{"type": r[0], "count": r[1]} for r in by_type],
        "top_vendors":[{"name": r[0], "count": r[1]} for r in top_vendors],
        "daily":      [{"date": r[0], "count": r[1]} for r in daily],
    }


# ── Excel export ───────────────────────────────────────────────────────────────

@app.get("/api/export/excel")
async def export_history_excel():
    con = get_db_connection()
    rows = con.execute(
        "SELECT id,processed_at,filename,invoice_type,invoice_number,"
        "invoice_date,party_name,total_amount,duplicate_of FROM invoices ORDER BY processed_at DESC"
    ).fetchall()
    con.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Invoice History"

    headers = ["ID","Processed At","Filename","Invoice Type","Invoice Number",
               "Invoice Date","Vendor / Party","Total Amount","Duplicate Of"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="CC1F1F")
        cell.alignment = Alignment(horizontal="center")

    for row in rows:
        ws.append(list(row))

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = min(
            max(len(str(c.value or "")) for c in col) + 4, 50
        )

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=samtel_invoices.xlsx"},
    )


@app.get("/api/history/{invoice_id}/excel")
async def export_invoice_excel(invoice_id: int):
    con = get_db_connection()
    row = con.execute("SELECT * FROM invoices WHERE id=?", (invoice_id,)).fetchone()
    con.close()
    if not row:
        raise HTTPException(404, "Invoice not found")

    # row: id,processed_at,filename,invoice_type,invoice_number,invoice_date,party_name,total_amount,data_json,duplicate_of
    filename     = row[2]
    invoice_type = row[3]
    data         = json.loads(row[8])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Invoice"

    # Title row
    ws.merge_cells("A1:B1")
    ws["A1"] = f"Samtel Avionics — {invoice_type.replace('_', ' ').title()}"
    ws["A1"].font  = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill  = PatternFill("solid", fgColor="CC1F1F")
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:B2")
    ws["A2"] = f"File: {filename}  |  Processed: {row[1][:10]}"
    ws["A2"].font = Font(italic=True, color="888888", size=9)
    ws["A2"].alignment = Alignment(horizontal="center")

    ws.append([])
    r = 4
    line_items = None

    for k, v in data.items():
        if k == "line_items" and isinstance(v, list):
            line_items = v
            continue
        if v is None:
            continue
        if isinstance(v, dict):
            ws.cell(r, 1, k.replace("_", " ").title())
            ws.cell(r, 1).font = Font(bold=True, color="CC1F1F")
            ws.merge_cells(f"A{r}:B{r}")
            r += 1
            for sk, sv in v.items():
                if sv is not None:
                    ws.cell(r, 1, "  " + sk.replace("_", " ").title()).font = Font(color="555555")
                    ws.cell(r, 2, str(sv))
                    r += 1
        else:
            ws.cell(r, 1, k.replace("_", " ").title()).font = Font(bold=True)
            ws.cell(r, 2, str(v))
            r += 1

    if line_items:
        r += 1
        ws.cell(r, 1, "Line Items").font = Font(bold=True, size=12, color="CC1F1F")
        r += 1
        if line_items:
            hdrs = list(line_items[0].keys())
            for j, h in enumerate(hdrs, 1):
                c = ws.cell(r, j, h.replace("_", " ").title())
                c.font = Font(bold=True, color="FFFFFF")
                c.fill = PatternFill("solid", fgColor="222222")
            r += 1
            for item in line_items:
                for j, h in enumerate(hdrs, 1):
                    ws.cell(r, j, str(item.get(h, "") or ""))
                r += 1

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 45

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe = filename.replace(" ", "_").replace("/", "_")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={safe}.xlsx"},
    )
# ── Static / Frontend ──────────────────────────────────────────────────────────

@app.get("/", response_class=HTMLResponse)
async def serve_frontend():
    html_path = Path(__file__).parent / "static" / "index.html"
    return html_path.read_text()


app.mount("/static", StaticFiles(directory="static"), name="static")

# Serve uploaded invoice files
uploads_dir = Path(__file__).parent / "uploads"
uploads_dir.mkdir(exist_ok=True)
app.mount("/uploads", StaticFiles(directory="uploads"), name="uploads")

